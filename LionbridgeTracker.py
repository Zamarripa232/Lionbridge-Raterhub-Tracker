#!/bin/env/ python
"""Automates time-keeping for Lionbridge's Raterhub job with output to an MS Excel spreadsheet.."""

__author__ = 'Floydasaurus'
__copyright__ = 'Copyright 2019, Floydasaur.us'

import datetime
import keyboard

import openpyxl  # pyWinhook, wx - might not need to use these?

# Initializing some variables/objects
currentAET = 1.0
filename = 'Data/Lionbridgetracker.xlsx'  # Put your own file here
startXKey = ']'
startSKey = '['
stopTimeKey = '\\'
exitKey = "~"  # Unlikely to press by accident.
activeRow = 2

wb = openpyxl.load_workbook(fileName)
activeSheet = wb.get_sheet_by_name('Times')


def getDate():
	"""Returns the date in the format mm/dd/yyyy"""
	return datetime.datetime.now().strftime('%m/%d/%Y')


def getTime():
	"""Returns the time in the format hh:mm:ss"""
	return datetime.datetime.now().strftime('%I:%M:%S')


def setAET(newAET):
	"""Sets the new estimated time, in minutes, for a Raterhub task."""
	currentAET = newAET
	return


def getFirstEmpty():
	for rowNum in range(2, activeSheet.max_row):
		emptyRowCheck = activeSheet.cell(row=rowNum, column=1).value
		if emptyRowCheck == "":
			return rowNum


def startxtimer():
	activeRow = getFirstEmpty()
	print(activeRow)
	activeSheet.cell(row=activeRow, column=1).value = getDate()
	activeSheet.cell(row=activeRow, column=2).value = getTime()
	activeSheet.cell(row=activeRow, column=3).value = 'X'
	activeSheet.cell(row=activeRow, column=4).value = currentAET
	return


def stoptimer():
	activeSheet.cell(row=activeRow, column=5).value = getTime()
	return


# Column Headers: A1Date | B2Start | C3Project | D4AET | E5End | F6Total Time

while True:
	if keyboard.is_pressed(startXKey):
		startxtimer()
		wb.save(filename)
	if keyboard.is_pressed(stopTimeKey):
		stoptimer()
		wb.save(filename)
	if keyboard.is_pressed(exitKey):
		break

wb.save(filename)